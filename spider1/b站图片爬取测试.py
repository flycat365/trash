# -*- coding: utf-8 -*-
#测试这里遇到的问题是search_box = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".nav-search-input")))
#search_button = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".nav-search-btn")))
#原版
#index = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#primary_menu > ul > li.home > a")))
#index.click()

# 等待搜索框元素出现
#input = WebDriverWait(browser, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#banner_link > div > div > form > input")))
#print(input.text)
#这里遇到的问题是b站的页面格式换了搜索框格式不对导致找不到
from bs4 import BeautifulSoup
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time

def get_source():
    # 获取当前页面的源代码
    source_code = browser.page_source
    return source_code


def find_span_tags(soup, class_name):
    # 找到具有指定类的span标签
    span_tags = soup.find_all('span', class_=class_name)

    # 遍历找到的span标签
    for tag in span_tags:
        # 打印出span标签的文本内容
        print(tag.get_text())

        # 如果span标签有子span标签，递归调用函数
        if tag.find('span'):
            find_span_tags(tag, class_name)

def parseHTML(html):
    """
    解析HTML页面
    """
    vlist=[]
    soup=BeautifulSoup(html, 'lxml')
    # 获取视频列表中的所有视频

    videos=soup.find_all('div', {'data-v-28552bfb': True})


    for video in videos:
        # 通过<a>标签获取视频名
        title=video.find('div', {'data-v-4caf9c8c': True}).get_text(strip=True)
        print(title)
        # 通过<a>标签获取视频链接
        href = video.find('a').get('href')[2:]
        print(href)
        # 通过<span>标签获取观看量
        times=video.find('span', {'data-v-4caf9c8c': True}).get_text(strip=True)

        # 通过<span>标签获取弹幕数

        print(title, href, times)
        vlist.append([title, href, times])

    return vlist
def Saver(vlist):
    """
    vlist：视频信息列表
    将数据保存在excel中
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '爬虫结果'
    sheet.cell(1,1,'视频名')
    sheet.cell(1,2,'视频链接')
    sheet.cell(1,3,'播放量')


    for idx,m in enumerate(vlist):
        sheet.cell(idx + 2,1,m[0])
        sheet.cell(idx + 2,2,m[1])
        sheet.cell(idx + 2,3,m[2])


    workbook.save('JianGuo_info.xlsx')  # 保存工作簿
# 初始化WebDriver，这里使用Edge浏览器
browser = webdriver.Edge()

# 打开哔哩哔哩首页
browser.get("https://www.bilibili.com/")

# 定义WebDriverWait实例
wait = WebDriverWait(browser, 10)

# 等待搜索框元素出现
search_box = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".nav-search-input")))

# 等待搜索按钮元素出现
search_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".nav-search-btn")))

# 在搜索框中输入内容
search_box.send_keys('京剧猫')
# 点击搜索按钮
search_button.click()

# 等待页面加载完成，这里使用一个示例条件，需要根据实际情况调整

# 跳转到新的窗口


# 定义翻页函数
def go_to_page(page_num):
    # 这里实现翻页逻辑，例如点击分页链接或按钮
    pass

# 尝试获取总页数
try:
    # 等待最后一页的分页按钮出现
    print('跳转到新窗口')
    all_h = browser.window_handles
    browser.switch_to.window(all_h[1])


    # 等待包含总页数信息的元素出现（需要根据实际页面结构调整选择器）<button class="vui_button vui_button--no-transition vui_pagenation--btn vui_pagenation--btn-num">34</button>
    #这里等待最下面的页数列出来，然后利用数组获得最后一个元素，然后获取总页数。
    pagination_num_buttons = wait.until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".vui_pagenation--btn.vui_pagenation--btn-num"))
    )

    # 从所有页码按钮中找到最后一个按钮，假设它是总页数
    # 注意：这里我们使用 reversed() 函数来反转按钮列表，然后取第一个元素
    total_pages_button = list(reversed(pagination_num_buttons))[0]

    # 获取总页数
    total_pages1= total_pages_button.text

    # 获取总页数，假设按钮的文本就是页码
    total_pages = int(total_pages_button.text)

    try:

        print(f"总页数为\n: {total_pages1}")
    except ValueError:
        print("无法解析总页数")
        total_pages = 1  # 如果解析失败，可以设置一个默认值，例如1
except TimeoutException:
    print("获取总页数失败，页面加载超时\n")
    browser.quit()
    exit()

# 翻页逻辑
try:
#<button class="vui_button vui_pagenation--btn vui_pagenation--btn-side">下一页</button>
    #for i in range(1, int(total_pages+ 1)):
    for i in range(1, 5):
        pagination_num_buttons = wait.until(
            EC.visibility_of_all_elements_located((By.CSS_SELECTOR, ".vui_pagenation--btn.vui_pagenation--btn-num"))
        )

        # 寻找下一页按钮，它可能是一个未被禁用的按钮，且不是当前页（不包含--active类）
        next_page_button =0
        for button in pagination_num_buttons:
            if 'vui_button--disabled' not in button.get_attribute(
                    'class') and 'vui_button--active' not in button.get_attribute('class'):
                next_page_button = button
                break

        # 如果找到下一页按钮，点击它
        if next_page_button:
            pagination_num_buttons = wait.until(
                EC.visibility_of_all_elements_located((By.CSS_SELECTOR, ".vui_pagenation--btn.vui_pagenation--btn-num"))
            )
            html = get_source()
            vlist = parseHTML(html)
            Saver(vlist)
            time.sleep(0.1)
            next_page_button.click()

            print(i)
        else:
            print("没有找到下一页按钮")
except TimeoutException:
    print("页面加载超时，尝试刷新页面")
    browser.refresh()

