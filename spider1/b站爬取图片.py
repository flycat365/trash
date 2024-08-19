# -*- coding: utf-8 -*-
#测试这里遇到的问题是search_box = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".nav-search-input")))
#search_button = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".nav-search-btn")))
#原版
#index = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#primary_menu > ul > li.home > a")))
#index.click()

# 等待搜索框元素出现
#input = WebDriverWait(browser, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#banner_link > div > div > form > input")))
#print(input.text)
#这里遇到的问题是b站的页面格式换了搜索框格式不对导致找不到
from bs4 import BeautifulSoup
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup  # 确保导入了 BeautifulSoup
import openpyxl

# ... [其他代码保持不变] ...
def get_source():
    # 获取当前页面的源代码
    source_code = browser.page_source
    return source_code
def save_to_excel(html):
    soup = BeautifulSoup(html, 'html.parser')
    video_list_row = soup.find('div', class_='video-list row')
    if video_list_row:
        list = video_list_row.find_all('div', class_='bili-video-card__wrap __scale-wrap')
    # ... [其他代码保持不变] ...
    for item in list:
        item_title = item.find('a').get('title') if item.find('a') else '标题未找到'
        item_link = item.find('a').get('href') if item.find('a') else '链接未找到'
        item_dec_element = item.find(class_='des hide')
        item_dec = item_dec_element.text if item_dec_element else '描述未找到'
        item_view = item.find('span', class_='bili-video-card__stats--item').text
        item_biubiu = item.find('span', class_='bili-video-card__stats--item').text
        item_date = item.find('span', class_='bili-video-card__stats__duration').text
        print(item_title, item_link, item_dec)
        print("\n")
        print(item_view, item_biubiu)
        print('爬取：' + item_title)
        n=0
        sheet.cell(n, 1, item_title)
        sheet.cell(n, 2, item_link)
        sheet.cell(n, 3, item_dec)
        sheet.cell(n, 4, item_view)
        sheet.cell(n, 5, item_biubiu)
        sheet.cell(n, 6, item_date)
        n += 1
    wb.save('output.xlsx')

browser = webdriver.Edge()

# 打开哔哩哔哩首页
browser.get("https://www.bilibili.com/")

# 定义WebDriverWait实例
wait = WebDriverWait(browser, 10)

# 等待搜索框元素出现
search_box = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".nav-search-input")))

# 等待搜索按钮元素出现
search_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".nav-search-btn")))

# 在搜索框中输入内容
search_box.send_keys('京剧猫')
# 点击搜索按钮
search_button.click()

# 等待页面加载完成，这里使用一个示例条件，需要根据实际情况调整

# 跳转到新的窗口


# 定义翻页函数
def go_to_page(page_num):
    # 这里实现翻页逻辑，例如点击分页链接或按钮
    pass

# 尝试获取总页数
try:
    # 等待最后一页的分页按钮出现
    print('跳转到新窗口')
    all_h = browser.window_handles
    browser.switch_to.window(all_h[1])


    # 等待包含总页数信息的元素出现（需要根据实际页面结构调整选择器）<button class="vui_button vui_button--no-transition vui_pagenation--btn vui_pagenation--btn-num">34</button>
    #这里等待最下面的页数列出来，然后利用数组获得最后一个元素，然后获取总页数。
    pagination_num_buttons = wait.until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".vui_pagenation--btn.vui_pagenation--btn-num"))
    )

    # 从所有页码按钮中找到最后一个按钮，假设它是总页数
    # 注意：这里我们使用 reversed() 函数来反转按钮列表，然后取第一个元素
    total_pages_button = list(reversed(pagination_num_buttons))[0]

    # 获取总页数
    total_pages1= total_pages_button.text

    # 获取总页数，假设按钮的文本就是页码
    total_pages = int(total_pages_button.text)

    try:

        print(f"总页数为\n: {total_pages1}")
    except ValueError:
        print("无法解析总页数")
        total_pages = 1  # 如果解析失败，可以设置一个默认值，例如1
except TimeoutException:
    print("获取总页数失败，页面加载超时\n")
    browser.quit()
    exit()

# 翻页逻辑
try:
#<button class="vui_button vui_pagenation--btn vui_pagenation--btn-side">下一页</button>
    #for i in range(1, int(total_pages+ 1)):
    for i in range(1, 5):
        pagination_num_buttons = wait.until(
            EC.visibility_of_all_elements_located((By.CSS_SELECTOR, ".vui_pagenation--btn.vui_pagenation--btn-num"))
        )

        # 寻找下一页按钮，它可能是一个未被禁用的按钮，且不是当前页（不包含--active类）
        next_page_button = None
        for button in pagination_num_buttons:
            if 'vui_button--disabled' not in button.get_attribute(
                    'class') and 'vui_button--active' not in button.get_attribute('class'):
                next_page_button = button
                break
        html = get_source()
        save_to_excel(html)
        # 如果找到下一页按钮，点击它
        if next_page_button:
            next_page_button.click()

            print(i)
        else:
            print("没有找到下一页按钮")
except TimeoutException:
    print("页面加载超时，尝试刷新页面")
    browser.refresh()


